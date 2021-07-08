#!/usr/bin/python
#-*- coding: utf-8 -*-
import json
import time

from netmiko import ConnectHandler
from openpyxl import Workbook

# 调用Openpyxl创建Workbook对象，Openpyxl还要很多很多诸如load_book, Color, PatternFill, Font, Border, Side等等之类的对象
# 它们的用法各不相同，一般我们在导入的时候只需要用from...import...来导入我们需要用到的对象就行了。
LogTime = time.strftime('%Y-%m-%d_%H-%M-%S')
device = {
        'device_type': 'cisco_ios',
        'host': '192.168.123.30',
        'username': 'cisco',
        'password': 'cisco',
 }

connect = ConnectHandler(**device)
print("已经成功登陆交换机:" + device['host'])
print("开始执行命令")
output = connect.send_command('show version',use_textfsm=True)
print(type(output))
log = open(device['host'] + '-' + LogTime + '.txt', 'w')
log.write(json.dumps(output,indent=2))


wb = Workbook()
ws = wb.active
ws.title = device['host']
ws['A1'] = 'Interfaces'
ws['B1'] = 'status'
ws['C1'] = 'Description'
#
row = 1;
for test in output:
        i = test.get('interface')
        s = test.get('link_status')
        print(i)
        row += 1
        print(row)
        ws.cell(row, column=1, value=i)
        ws.cell(row, column=2, value=s)

wb.save(device['host']+ LogTime +'.xlsx')           # 保存表格