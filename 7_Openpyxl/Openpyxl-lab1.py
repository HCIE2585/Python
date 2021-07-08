from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
# 调用Openpyxl创建Workbook对象，Openpyxl还要很多很多诸如load_book, Color, PatternFill, Font, Border, Side等等之类的对象
# 它们的用法各不相同，一般我们在导入的时候只需要用from...import...来导入我们需要用到的对象就行了。
wb = Workbook()                                         # 用来创建一个workbook，并将它赋值给变量wb
# 操作sheet名称和位置
ws = wb.active                                          # 用wb.active生成第一个工作表(worksheet)，赋值给变量ws，该工作表默认名为'Sheet'
ws.title = 'Router'                                     # 要将该工作表名从Sheet改成Router
ws1 = wb.create_sheet("Switch", 1)                      # 创建一个新sheet，赋值给变量ws1，sheet名字是 Switch，参数1表示将该工作表放在所有工作表中的第2位。不带参数放在最后
ws2 = wb.create_sheet("Firewall", 2)
print('wb.sheetnames来查看所有工作表的名字')
print(wb.sheetnames)                                    # wb.sheetnames来查看所有工作表的名字
print('用for循环来查看sheet名字')
for sheet in wb:                                        # 也可以用for循环来查看sheet名字
        print(sheet.title)
# 修改单元格内容
ws1['A1'] = 'Number'                                    # 通过ws['单元格号']的形式来为指定的单元格添加内容
ws1['B1'] = 'interface'
ws1['C1'] = 'description'
ws1.cell(row=2, column=2, value='Gi1/0/1')              # 使用ws.cell()函数来修改单元格的内容,其中参数row代表排数，column代表列数，value代表要写入单元格的内容
ws1.cell(row=3, column=2, value='Gi1/0/2')
ws1.cell(row=2, column=3, value='PC1')
ws1.cell(row=3, column=3, value='PC2')

print('查看具体某个单元格内容')
b2 =  ws1['B2']
print(b2.value)                                         # 使用value属性查看单元格内容

print('查看具体一列单元格内容')
column_b = ws1['B']
for i in column_b:
        print(i.value)

print('查看具体一行单元格内容')
row_2 =  ws1['2']
for i in row_2:
        print(i.value)

print('查看ws1中的A1-A3,C1-C3内容')
for row in ws1.iter_rows(min_row=1, max_col=3, max_row=3):
        for cell in row:
                print (cell.value)                      # 显示顺序是按行显示

print('查看ws1中的所有内容')
for row in ws1.values:
        for value in row:
                print(value)

print('修改单元格内容')
b2 =  ws1['B2']
print('修改前的单元格内容为：' + b2.value)
b2.value = 'Gi1/0/100ssssssssssssss'
print('修改后的单元格内容为：' + b2.value)


wb.save('Openpyxl-lab1.xlsx')                           # 保存表格
