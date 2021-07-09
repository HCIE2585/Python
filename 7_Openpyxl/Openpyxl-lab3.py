#!/usr/bin/python
# -*- coding: utf-8 -*-
import json
import time
from netmiko import ConnectHandler
from openpyxl import Workbook
LogTime = time.strftime('%Y-%m-%d_%H-%M-%S')
with open('Openpyxl-lab3.txt') as f:
    wb = Workbook()
    ws = wb.active
    ws.title = 'System-Info'  # 默认sheet为设备系统信息
    ws['A1'] = 'Number'
    ws['B1'] = 'Hostname'
    ws['C1'] = 'Hardware Platform'
    ws['D1'] = 'Software Version'
    ws['E1'] = '系统运行时间'
    ws['F1'] = 'Serial Number'
    ws1 = wb.create_sheet('Cpu-Info')  # 创建sheet，名称为Cpu-Info
    ws1['A1'] = 'Number'
    ws1['B1'] = 'Hostname'
    ws1['C1'] = 'cpu_5_sec'
    ws1['D1'] = 'cpu_1_min'
    ws1['E1'] = 'cpu_5_min'
    ws2 = wb.create_sheet('Memory-Info')  # 创建sheet，名称为Memory-Info
    ws2['A1'] = 'Number'
    ws2['B1'] = 'Hostname'
    ws2['C1'] = '内容总容量'
    ws2['D1'] = '已使用'
    ws2['E1'] = '未使用'
    ws2['F1'] = '使用率'
    sys_row = 1
    cpu_row = 1
    mem_row = 1
    for ips in f.readlines():
        ip = ips.strip()
        device = {
            'device_type': 'cisco_ios',
            'host': ip,
            'username': 'cisco',
            'password': 'cisco',
        }
        try:
            connect = ConnectHandler(**device)
            print('Start to connect:' + device['host'])
            sys_info = connect.send_command('show version', use_textfsm=True)
            cpu_info = connect.send_command('show processes cpu', use_textfsm=True)
            memo_info = connect.send_command('show processes memory', use_genie=True)
            print(json.dumps(memo_info, indent=2))
            intf = connect.send_command('show interface', use_textfsm=True)
            inventory = connect.send_command('show inventory', use_genie=True)
            log_info = connect.send_command('show logging', use_genie=True)
            log = open(LogTime + '.txt', 'w')
            log.write(json.dumps(log_info, indent=2))
            log.close()
            for sys in sys_info:  # 读取设备系统相关信息
                hostname = sys.get('hostname')
                hardware = sys.get('hardware')
                verison = sys.get('version')
                uptime = sys.get('uptime')
                serial = sys.get('serial')
                sys_row += 1
                ws.cell(sys_row, column=1, value=(sys_row - 1))
                ws.cell(sys_row, column=2, value=hostname)
                ws.cell(sys_row, column=3, value=hardware[0])
                ws.cell(sys_row, column=4, value=verison)
                ws.cell(sys_row, column=5, value=uptime)
                ws.cell(sys_row, column=6, value=serial[0])
                print('sys')
            for cpu in cpu_info:  # 读取设备cpu相关信息
                hostname = sys.get('hostname')
                sec5 = cpu.get('cpu_5_sec')
                min1 = cpu.get('cpu_1_min')
                min5 = cpu.get('cpu_5_min')
                cpu_row += 1
                ws1.cell(cpu_row, column=1, value=(cpu_row - 1))
                ws1.cell(cpu_row, column=2, value=hostname)
                ws1.cell(cpu_row, column=3, value=sec5)
                ws1.cell(cpu_row, column=4, value=min1)
                ws1.cell(cpu_row, column=5, value=min5)
                print('cpu')
            mem = memo_info['processor_pool']  # 读取设备内存相关信息
            hostname = sys.get('hostname')
            total = mem.get('total')
            used = mem.get('used')
            free = mem.get('free')
            mem_row += 1
            ws2.cell(mem_row, column=1, value=(mem_row - 1))
            ws2.cell(mem_row, column=2, value=hostname)
            ws2.cell(mem_row, column=3, value=total)
            ws2.cell(mem_row, column=4, value=used)
            ws2.cell(mem_row, column=5, value=free)
            ws2.cell(mem_row, column=6, value=used / total)
            sheet = 100
            ws.sheet = wb.create_sheet(device['host'] + '-Interface-Info')  # 创建接口信息sheet，一个设备一个sheet
            ws.sheet['A1'] = 'Number'
            ws.sheet['B1'] = 'Interface'
            ws.sheet['C1'] = 'link_status'
            ws.sheet['D1'] = 'ip_address'
            ws.sheet['E1'] = 'Description'
            int_row = 1
            for info in intf:  # 读取设备接口输出信息
                interface = info.get('interface')
                status = info.get('link_status')
                ip_addr = info.get('ip_address')
                description = info.get('description')
                int_row += 1
                ws.sheet.cell(int_row, column=1, value=(int_row - 1))
                ws.sheet.cell(int_row, column=2, value=interface)
                ws.sheet.cell(int_row, column=3, value=status)
                ws.sheet.cell(int_row, column=4, value=ip_addr)
                ws.sheet.cell(int_row, column=5, value=description)
        except Exception as e:
            print(f'connection error ip:{ip} error:{str(e)}')
    wb.save(LogTime + '.xlsx')
